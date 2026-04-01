#!/usr/bin/env python3
"""
GCP Discovery Scanner — Phase 1 Automation
Scans the GCP organization to inventory all existing resources,
projects, IAM bindings, networks, and compliance posture.

Usage:
    python -m scripts.python.discovery.scanner --org-id 123456789 --output json
    python -m scripts.python.discovery.scanner --org-id 123456789 --output xlsx --file discovery.xlsx
"""

import argparse
import json
import logging
import subprocess
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("gcp-discovery")


@dataclass
class GCPProject:
    project_id: str
    name: str
    state: str
    parent_type: str
    parent_id: str
    labels: dict = field(default_factory=dict)
    create_time: str = ""
    billing_enabled: bool = False
    apis_enabled: list = field(default_factory=list)
    iam_bindings: list = field(default_factory=list)
    networks: list = field(default_factory=list)


@dataclass
class DiscoveryReport:
    org_id: str
    scan_time: str
    projects: list = field(default_factory=list)
    org_policies: list = field(default_factory=list)
    iam_summary: dict = field(default_factory=dict)
    network_summary: dict = field(default_factory=dict)
    compliance_gaps: list = field(default_factory=list)
    recommendations: list = field(default_factory=list)


def run_gcloud(args: list[str], fmt: str = "json") -> Optional[dict | list]:
    """Execute a gcloud command and return parsed output."""
    cmd = ["gcloud"] + args + [f"--format={fmt}"]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode != 0:
            log.warning(f"gcloud command failed: {' '.join(cmd)}\n{result.stderr}")
            return None
        if fmt == "json" and result.stdout.strip():
            return json.loads(result.stdout)
        return result.stdout.strip()
    except subprocess.TimeoutExpired:
        log.error(f"Command timed out: {' '.join(cmd)}")
        return None
    except json.JSONDecodeError:
        log.error(f"Failed to parse JSON from: {' '.join(cmd)}")
        return None


def discover_projects(org_id: str) -> list[GCPProject]:
    """Enumerate all projects under the organization."""
    log.info(f"Discovering projects under org {org_id}...")
    raw = run_gcloud([
        "projects", "list",
        f"--filter=parent.id={org_id}",
    ])
    if not raw:
        return []

    projects = []
    for p in raw:
        proj = GCPProject(
            project_id=p.get("projectId", ""),
            name=p.get("name", ""),
            state=p.get("lifecycleState", "UNKNOWN"),
            parent_type=p.get("parent", {}).get("type", ""),
            parent_id=p.get("parent", {}).get("id", ""),
            labels=p.get("labels", {}),
            create_time=p.get("createTime", ""),
        )
        projects.append(proj)

    log.info(f"  Found {len(projects)} projects")
    return projects


def discover_project_details(project: GCPProject) -> GCPProject:
    """Enrich a project with IAM, APIs, and network data."""
    log.info(f"  Scanning project: {project.project_id}")

    # Check billing
    billing = run_gcloud([
        "billing", "projects", "describe", project.project_id
    ])
    if billing and isinstance(billing, dict):
        project.billing_enabled = billing.get("billingEnabled", False)

    # List enabled APIs
    apis = run_gcloud([
        "services", "list",
        f"--project={project.project_id}",
        "--enabled",
    ])
    if apis:
        project.apis_enabled = [a.get("config", {}).get("name", "") for a in apis]

    # IAM policy
    iam = run_gcloud([
        "projects", "get-iam-policy", project.project_id,
    ])
    if iam and isinstance(iam, dict):
        project.iam_bindings = iam.get("bindings", [])

    # Networks
    networks = run_gcloud([
        "compute", "networks", "list",
        f"--project={project.project_id}",
    ])
    if networks:
        project.networks = [{"name": n.get("name"), "mode": n.get("autoCreateSubnetworks")} for n in networks]

    return project


def discover_org_policies(org_id: str) -> list[dict]:
    """List all org policy constraints and their enforcement status."""
    log.info("Discovering organization policies...")
    constraints = [
        "compute.requireShieldedVm",
        "compute.vmExternalIpAccess",
        "iam.disableServiceAccountKeyCreation",
        "iam.disableServiceAccountKeyUpload",
        "storage.publicAccessPrevention",
        "storage.uniformBucketLevelAccess",
        "compute.skipDefaultNetworkCreation",
        "compute.requireOsLogin",
        "gcp.resourceLocations",
        "iam.allowedPolicyMemberDomains",
    ]

    policies = []
    for c in constraints:
        result = run_gcloud([
            "resource-manager", "org-policies", "describe",
            f"constraints/{c}",
            f"--organization={org_id}",
        ])
        status = "UNKNOWN"
        if result and isinstance(result, dict):
            if result.get("booleanPolicy", {}).get("enforced"):
                status = "ENFORCED"
            elif result.get("listPolicy"):
                status = "LIST_ACTIVE"
            else:
                status = "NOT_ENFORCED"
        policies.append({"constraint": c, "status": status})

    enforced = sum(1 for p in policies if p["status"] in ("ENFORCED", "LIST_ACTIVE"))
    log.info(f"  {enforced}/{len(policies)} org policies enforced")
    return policies


def analyze_compliance_gaps(report: DiscoveryReport) -> list[dict]:
    """Identify HIPAA compliance gaps from discovery data."""
    gaps = []

    # Check for default networks
    for proj in report.projects:
        for net in proj.networks:
            if net.get("name") == "default":
                gaps.append({
                    "severity": "HIGH",
                    "project": proj.project_id,
                    "finding": "Default VPC network exists",
                    "remediation": "Delete default network; enforce compute.skipDefaultNetworkCreation",
                    "hipaa_control": "§164.312(e)(1) — Transmission Security",
                })

    # Check for SA key creation policy
    sa_key_policy = next((p for p in report.org_policies if p["constraint"] == "iam.disableServiceAccountKeyCreation"), None)
    if sa_key_policy and sa_key_policy["status"] != "ENFORCED":
        gaps.append({
            "severity": "CRITICAL",
            "finding": "SA key creation not disabled",
            "remediation": "Enforce iam.disableServiceAccountKeyCreation; migrate to WIF (ADR-015)",
            "hipaa_control": "§164.312(d) — Person or Entity Authentication",
        })

    # Check for public access prevention
    pub_policy = next((p for p in report.org_policies if p["constraint"] == "storage.publicAccessPrevention"), None)
    if pub_policy and pub_policy["status"] != "ENFORCED":
        gaps.append({
            "severity": "CRITICAL",
            "finding": "GCS public access prevention not enforced",
            "remediation": "Enforce storage.publicAccessPrevention org-wide",
            "hipaa_control": "§164.312(a)(1) — Access Control",
        })

    # Check for projects without labels
    for proj in report.projects:
        if not proj.labels or "environment" not in proj.labels:
            gaps.append({
                "severity": "MEDIUM",
                "project": proj.project_id,
                "finding": "Project missing environment label",
                "remediation": "Apply standard labels: environment, compliance, managed_by",
                "hipaa_control": "§164.310(d)(1) — Device and Media Controls",
            })

    return gaps


def generate_recommendations(report: DiscoveryReport) -> list[dict]:
    """Generate actionable recommendations from the discovery scan."""
    recs = []

    # Count projects by state
    active = sum(1 for p in report.projects if p.state == "ACTIVE")
    if active > 20:
        recs.append({
            "priority": "HIGH",
            "category": "Governance",
            "recommendation": f"Consolidate {active} active projects into folder hierarchy per ADR-003",
            "effort": "Medium",
        })

    # Check for unused APIs
    for proj in report.projects:
        risky_apis = [a for a in proj.apis_enabled if a in (
            "deploymentmanager.googleapis.com",
            "appengine.googleapis.com",
        )]
        if risky_apis:
            recs.append({
                "priority": "MEDIUM",
                "category": "Security",
                "recommendation": f"Disable unused APIs in {proj.project_id}: {', '.join(risky_apis)}",
                "effort": "Low",
            })

    # Compliance gaps become recommendations
    critical_gaps = [g for g in report.compliance_gaps if g["severity"] == "CRITICAL"]
    if critical_gaps:
        recs.append({
            "priority": "CRITICAL",
            "category": "Compliance",
            "recommendation": f"Address {len(critical_gaps)} critical HIPAA compliance gaps before migration",
            "effort": "High",
        })

    return recs


def export_json(report: DiscoveryReport, path: str):
    """Export discovery report as JSON."""
    data = {
        "org_id": report.org_id,
        "scan_time": report.scan_time,
        "projects": [asdict(p) for p in report.projects],
        "org_policies": report.org_policies,
        "compliance_gaps": report.compliance_gaps,
        "recommendations": report.recommendations,
        "summary": {
            "total_projects": len(report.projects),
            "active_projects": sum(1 for p in report.projects if p.state == "ACTIVE"),
            "policies_enforced": sum(1 for p in report.org_policies if p["status"] == "ENFORCED"),
            "critical_gaps": sum(1 for g in report.compliance_gaps if g["severity"] == "CRITICAL"),
        },
    }
    Path(path).write_text(json.dumps(data, indent=2, default=str))
    log.info(f"Report exported to {path}")


def export_xlsx(report: DiscoveryReport, path: str):
    """Export discovery report as XLSX workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.error("openpyxl required for XLSX export: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # Projects sheet
    ws = wb.active
    ws.title = "Projects"
    headers = ["Project ID", "Name", "State", "Parent", "Billing", "APIs", "Networks", "Labels"]
    ws.append(headers)
    for h in range(1, len(headers) + 1):
        ws.cell(1, h).font = Font(bold=True)
        ws.cell(1, h).fill = PatternFill("solid", fgColor="1a73e8")
        ws.cell(1, h).font = Font(bold=True, color="FFFFFF")

    for proj in report.projects:
        ws.append([
            proj.project_id, proj.name, proj.state,
            f"{proj.parent_type}/{proj.parent_id}",
            "Yes" if proj.billing_enabled else "No",
            len(proj.apis_enabled),
            len(proj.networks),
            json.dumps(proj.labels),
        ])

    # Compliance Gaps sheet
    ws2 = wb.create_sheet("Compliance Gaps")
    ws2.append(["Severity", "Project", "Finding", "HIPAA Control", "Remediation"])
    for h in range(1, 6):
        ws2.cell(1, h).font = Font(bold=True, color="FFFFFF")
        ws2.cell(1, h).fill = PatternFill("solid", fgColor="d93025")
    for gap in report.compliance_gaps:
        ws2.append([gap.get("severity"), gap.get("project", "Org-wide"), gap["finding"], gap.get("hipaa_control", ""), gap["remediation"]])

    # Org Policies sheet
    ws3 = wb.create_sheet("Org Policies")
    ws3.append(["Constraint", "Status"])
    for h in range(1, 3):
        ws3.cell(1, h).font = Font(bold=True, color="FFFFFF")
        ws3.cell(1, h).fill = PatternFill("solid", fgColor="0d652d")
    for pol in report.org_policies:
        ws3.append([pol["constraint"], pol["status"]])

    # Recommendations sheet
    ws4 = wb.create_sheet("Recommendations")
    ws4.append(["Priority", "Category", "Recommendation", "Effort"])
    for h in range(1, 5):
        ws4.cell(1, h).font = Font(bold=True, color="FFFFFF")
        ws4.cell(1, h).fill = PatternFill("solid", fgColor="f9ab00")
    for rec in report.recommendations:
        ws4.append([rec["priority"], rec["category"], rec["recommendation"], rec["effort"]])

    wb.save(path)
    log.info(f"XLSX report exported to {path}")


def main():
    parser = argparse.ArgumentParser(description="GCP Discovery Scanner")
    parser.add_argument("--org-id", required=True, help="GCP Organization ID")
    parser.add_argument("--output", choices=["json", "xlsx", "both"], default="json")
    parser.add_argument("--file", default="discovery-report", help="Output filename (no extension)")
    parser.add_argument("--deep", action="store_true", help="Deep scan: enumerate IAM, APIs, networks per project")
    args = parser.parse_args()

    log.info("╔═══════════════════════════════════════════════════════╗")
    log.info("║  GCP Chaotic Deploy — Discovery Scanner v1.0         ║")
    log.info("╚═══════════════════════════════════════════════════════╝")

    report = DiscoveryReport(
        org_id=args.org_id,
        scan_time=datetime.utcnow().isoformat() + "Z",
    )

    # Phase 1: Discover projects
    report.projects = discover_projects(args.org_id)

    # Phase 2: Deep scan (optional)
    if args.deep:
        report.projects = [discover_project_details(p) for p in report.projects]

    # Phase 3: Org policies
    report.org_policies = discover_org_policies(args.org_id)

    # Phase 4: Compliance analysis
    report.compliance_gaps = analyze_compliance_gaps(report)
    report.recommendations = generate_recommendations(report)

    # Export
    if args.output in ("json", "both"):
        export_json(report, f"{args.file}.json")
    if args.output in ("xlsx", "both"):
        export_xlsx(report, f"{args.file}.xlsx")

    log.info(f"\n{'═'*55}")
    log.info(f"  Projects discovered : {len(report.projects)}")
    log.info(f"  Org policies checked: {len(report.org_policies)}")
    log.info(f"  Compliance gaps     : {len(report.compliance_gaps)}")
    log.info(f"  Recommendations     : {len(report.recommendations)}")
    log.info(f"{'═'*55}")


if __name__ == "__main__":
    main()
